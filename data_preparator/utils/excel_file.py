import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from ..exceptions import MissingColumnsInDataFrameError


def convert_data_frame_to_workbook_of_openpyxl(df: pd.DataFrame) -> Workbook:
    """Конвертирует дата фрейм в воркбук пакета openpyxl."""
    wb = Workbook()
    ws = wb.active
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    return wb


def apply_style(wb: Workbook) -> None:
    """Применяет excel стили."""
    ws = wb.active
    ws.merge_cells('A2:I6')
    cell_with_error_message = ws['A2']
    cell_with_error_message.alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True,
    )
    cell_with_error_message.border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick'),
    )
    cell_with_error_message.fill = PatternFill(
        start_color='FFEFEF',
        end_color='FFEFEF',
        fill_type='solid',
    )


def insert_rows_with_file_errors_into_workbook(wb: Workbook, error: MissingColumnsInDataFrameError) -> None:
    """Вставляет пустые строки (потом они будут смёрджены) и в ячейку А2 текст ошибок в эксель таблицу."""
    ws = wb.active
    ws.insert_rows(2, amount=5)
    ws['A2'] = error.message
