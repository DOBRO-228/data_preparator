import itertools

import pandas as pd
from data_preparator.exceptions import MissingColumnsInDataFrameError
from data_preparator.utils.excel_file import (
    convert_data_frame_to_workbook_of_openpyxl,
    insert_rows_with_file_errors_into_workbook,
)
from openpyxl import Workbook


def test_convert_data_frame_to_workbook_of_openpyxl():
    df_data = {'GACHI': ['Billy', 'Dungeon Master']}
    df = pd.DataFrame(
        data=df_data,
    )
    df_as_workbook = convert_data_frame_to_workbook_of_openpyxl(df)

    assert isinstance(df_as_workbook, Workbook)

    ws = df_as_workbook.active
    expected_values_in_column = list(itertools.chain(
        df_data.keys(), *df_data.values()
    ))
    for cell in ws['A']:
        assert cell.value in expected_values_in_column


def test_insert_rows_with_file_errors_into_workbook():
    errors = MissingColumnsInDataFrameError(['3', 'hundred', 'bucks'])
    wb = Workbook()
    insert_rows_with_file_errors_into_workbook(wb, errors)
    ws = wb.active
    assert ws['A2'].value == errors.message
